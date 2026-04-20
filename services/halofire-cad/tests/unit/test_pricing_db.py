"""Unit + smoke tests for the supplies pricing DB.

Covers:
  - idempotent schema bootstrap
  - append-only prices (no silent overwrites)
  - `price_for` returns latest
  - stale flag flips at STALE_DAYS
  - apply_updates rejects unknown SKU
  - Excel export has expected sheets
  - sync_runs audit trail
  - CSV deterministic sync path

Run:
  pytest services/halofire-cad/tests/unit/test_pricing_db.py -q
"""
from __future__ import annotations

import csv
import importlib.util
import sys
from datetime import datetime, timedelta
from pathlib import Path

import pytest

_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(_ROOT))

from pricing.db import (  # noqa: E402
    STALE_DAYS, PriceUpdate, SuppliesDB, SyncRun,
    open_db, sha256_of, utcnow_naive,
)


@pytest.fixture
def db(tmp_path: Path) -> SuppliesDB:
    d = SuppliesDB(tmp_path / "supplies.duckdb")
    d.upsert_supplier("testmfr", "Test MFR")
    d.upsert_part(
        {
            "sku": "TEST-1", "name": "Test part", "category": "fitting_elbow_90",
            "mounting": "pipe_inline", "manufacturer": "Test MFR",
            "supplier_id": "testmfr", "model": "T1",
            "dim_l_cm": None, "dim_d_cm": None, "dim_h_cm": None,
            "pipe_size_in": 2.0, "k_factor": None, "temp_rating_f": None,
            "response": None, "connection": "grooved", "finish": "Black iron",
            "nfpa_paint_hex": None, "open_source_glb": False,
            "discontinued": False, "notes": "",
        },
    )
    yield d
    d.close()


def test_bootstrap_idempotent(tmp_path: Path) -> None:
    """Opening the same file twice must not raise."""
    p = tmp_path / "s.duckdb"
    SuppliesDB(p).close()
    d = SuppliesDB(p)  # bootstrap runs again
    assert d.part_count() == 0
    d.close()


def test_price_append_only(db: SuppliesDB) -> None:
    db.add_price(PriceUpdate(sku="TEST-1", unit_cost_usd=10.00, source="t1"))
    db.add_price(PriceUpdate(sku="TEST-1", unit_cost_usd=12.00, source="t2"))
    rows = db.con().execute(
        "SELECT COUNT(*) FROM prices WHERE sku='TEST-1'",
    ).fetchone()
    assert rows[0] == 2


def test_price_for_returns_latest(db: SuppliesDB) -> None:
    # Insert old then new; `price_for` must return new
    old = utcnow_naive() - timedelta(days=30)
    db.con().execute(
        "INSERT INTO prices (sku, unit_cost_usd, unit, observed_at, source, confidence, currency) "
        "VALUES ('TEST-1', 5.0, 'ea', ?, 'old', 1.0, 'USD')",
        [old],
    )
    db.add_price(PriceUpdate(sku="TEST-1", unit_cost_usd=11.00, source="new"))
    row = db.price_for("TEST-1")
    assert row is not None
    assert row.unit_cost_usd == 11.00
    assert row.source == "new"
    assert row.stale is False


def test_stale_flag_crosses_threshold(db: SuppliesDB) -> None:
    old = utcnow_naive() - timedelta(days=STALE_DAYS + 5)
    db.con().execute(
        "INSERT INTO prices (sku, unit_cost_usd, unit, observed_at, source, confidence, currency) "
        "VALUES ('TEST-1', 5.0, 'ea', ?, 'ancient', 1.0, 'USD')",
        [old],
    )
    row = db.price_for("TEST-1")
    assert row is not None
    assert row.stale is True


def test_price_for_missing_sku_returns_none(db: SuppliesDB) -> None:
    assert db.price_for("DOES-NOT-EXIST") is None


def test_apply_updates_rejects_unknown_sku(db: SuppliesDB) -> None:
    accepted, errs = db.apply_updates(
        [
            PriceUpdate(sku="NOT-A-PART", unit_cost_usd=5.0, source="test"),
            PriceUpdate(sku="TEST-1", unit_cost_usd=9.0, source="test"),
        ],
    )
    assert accepted == 1
    assert any("NOT-A-PART" in e for e in errs)


def test_price_update_validate_catches_bad_unit() -> None:
    u = PriceUpdate(sku="x", unit_cost_usd=5.0, unit="barrel")
    assert "unit" in " ".join(u.validate())


def test_price_update_validate_catches_negative_cost() -> None:
    u = PriceUpdate(sku="x", unit_cost_usd=-0.01)
    assert "unit_cost_usd" in " ".join(u.validate())


def test_price_update_validate_catches_bad_confidence() -> None:
    u = PriceUpdate(sku="x", unit_cost_usd=1.0, confidence=1.5)
    assert "confidence" in " ".join(u.validate())


def test_stale_skus_view_surfaces_gaps(db: SuppliesDB) -> None:
    # TEST-1 has no price at all — must appear in stale_skus
    stale = db.stale_skus()
    assert any(sku == "TEST-1" for sku, _ in stale)
    # After adding a current price, it must NOT appear
    db.add_price(PriceUpdate(sku="TEST-1", unit_cost_usd=5.0, source="t"))
    fresh = db.stale_skus()
    assert not any(sku == "TEST-1" for sku, _ in fresh)


def test_sync_run_round_trip(db: SuppliesDB) -> None:
    run = SyncRun(supplier_id="testmfr", source_url="x", llm_model="none")
    run_id = db.start_sync_run(run)
    db.finish_sync_run(run_id, parts_touched=3, prices_added=2, status="success")
    recent = db.recent_sync_runs(limit=5)
    assert recent
    latest = recent[0]
    assert latest["id"] == run_id
    assert latest["prices_added"] == 2
    assert latest["status"] == "success"


def test_export_xlsx_has_all_sheets(db: SuppliesDB, tmp_path: Path) -> None:
    db.add_price(PriceUpdate(sku="TEST-1", unit_cost_usd=5.0, source="t"))
    out = db.export_xlsx(tmp_path / "snap.xlsx")
    assert out.exists() and out.stat().st_size > 1000
    from openpyxl import load_workbook
    wb = load_workbook(out)
    # parts + latest_prices + sync_runs all present
    assert {"parts", "latest_prices", "sync_runs"}.issubset(set(wb.sheetnames))


def test_sha256_deterministic(tmp_path: Path) -> None:
    p = tmp_path / "f.txt"
    p.write_text("halofire", encoding="utf-8")
    assert sha256_of(p) == sha256_of(p)
    assert len(sha256_of(p)) == 64


# ── sync agent (CSV path — no LLM required) ──────────────────────

def test_sync_agent_csv_round_trip(tmp_path: Path, db: SuppliesDB) -> None:
    """CSV path is fully deterministic — no Ollama needed."""
    sa_spec = importlib.util.spec_from_file_location(
        "_sync_agent",
        _ROOT / "pricing" / "sync_agent.py",
    )
    assert sa_spec is not None and sa_spec.loader is not None
    sa = importlib.util.module_from_spec(sa_spec)
    sa_spec.loader.exec_module(sa)

    # Write a CSV — real supplier feed shape
    csv_path = tmp_path / "feed.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sku", "unit_cost_usd", "unit"])
        w.writerow(["TEST-1", "7.25", "ea"])
        w.writerow(["NOT-A-PART", "1.00", "ea"])  # rejected

    # Monkey-patch the sync_agent to use our test DB
    import contextlib

    orig_open = sa.open_db

    @contextlib.contextmanager
    def _test_open(*_a, **_kw):
        yield db

    sa.open_db = _test_open
    try:
        res = sa.run_sync(
            supplier_id="testmfr", source_path=csv_path, dry_run=False,
        )
    finally:
        sa.open_db = orig_open

    assert res["accepted"] == 1
    assert any("NOT-A-PART" in e for e in res["errors"])
    row = db.price_for("TEST-1")
    assert row is not None and row.unit_cost_usd == 7.25
    # Source string carries the supplier id + short hash for audit
    assert row.source.startswith("sync_agent:testmfr:")


def test_gemma_only_policy_rejects_non_gemma_tag() -> None:
    """sync_agent._require_gemma is the code-level enforcement of
    the 'Gemma only' rule. Any non-Gemma tag (Qwen, Llama, Mistral,
    empty string, random garbage) must raise ValueError BEFORE the
    agent touches Ollama, the DB, or the network."""
    sa_spec = importlib.util.spec_from_file_location(
        "_sa_guard", _ROOT / "pricing" / "sync_agent.py",
    )
    assert sa_spec is not None and sa_spec.loader is not None
    sa = importlib.util.module_from_spec(sa_spec)
    sa_spec.loader.exec_module(sa)

    # Accepted tags — anything starting with 'gemma', 'gemma2', 'gemma3'
    for ok in ("gemma3:4b", "gemma3:12b", "gemma2:9b", "Gemma3:27b"):
        sa._require_gemma(ok)  # must not raise

    # Rejected: every non-Gemma family
    for bad in ("qwen2.5:7b", "qwen3:8b", "llama3:8b",
                "mistral:7b", "phi3:mini", "", "random"):
        with pytest.raises(ValueError, match="Gemma-only"):
            sa._require_gemma(bad)


def test_default_model_is_gemma() -> None:
    """The exported DEFAULT_MODEL must be a Gemma tag — ensures a
    misconfigured env var can't silently flip us back to Qwen."""
    sa_spec = importlib.util.spec_from_file_location(
        "_sa_default", _ROOT / "pricing" / "sync_agent.py",
    )
    assert sa_spec is not None and sa_spec.loader is not None
    sa = importlib.util.module_from_spec(sa_spec)
    sa_spec.loader.exec_module(sa)
    assert sa.DEFAULT_MODEL.lower().startswith("gemma")


def test_sync_run_dry_run_commits_nothing(tmp_path: Path, db: SuppliesDB) -> None:
    sa_spec = importlib.util.spec_from_file_location(
        "_sa2", _ROOT / "pricing" / "sync_agent.py",
    )
    assert sa_spec is not None and sa_spec.loader is not None
    sa = importlib.util.module_from_spec(sa_spec)
    sa_spec.loader.exec_module(sa)

    import contextlib

    @contextlib.contextmanager
    def _test_open(*_a, **_kw):
        yield db

    sa.open_db = _test_open
    csv_path = tmp_path / "feed.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sku", "unit_cost_usd", "unit"])
        w.writerow(["TEST-1", "42.0", "ea"])

    res = sa.run_sync(
        supplier_id="testmfr", source_path=csv_path, dry_run=True,
    )
    assert res["accepted"] == 0
    # The price_for should still be whatever it was before (None here)
    assert db.price_for("TEST-1") is None
