"""halofire proposal agent — produces the web-bid data + PDF + XLSX.

Inputs:
  - Design (classified, placed, routed, calc'd, rule-checked)
  - BOM rows
  - Labor rows

Outputs (all saved to project's deliverables directory):
  - proposal.json   — structured data the web bid viewer consumes
  - proposal.pdf    — PDF proposal matching Halo's format
  - proposal.xlsx   — Halo's pricing workbook (for internal P&L tracking)

The proposal.json is the CANONICAL output — web bid viewer on both
desktop and mobile reads it. PDF + XLSX are derived.
"""
from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from cad.schema import Design, BomRow, LaborRow  # noqa: E402


def _level_summary(design: Design) -> list[dict]:
    """One row per level with head + pipe + hazard counts."""
    out = []
    for level in design.building.levels:
        heads_here = [
            h for s in design.systems for h in s.heads
            if any(r.id == h.room_id for r in level.rooms)
        ]
        pipes_here = [
            p for s in design.systems for p in s.pipes
            if p.system_id and level.id in s.supplies
        ]
        out.append({
            "id": level.id,
            "name": level.name,
            "use": level.use,
            "elevation_m": level.elevation_m,
            "elevation_ft": round(level.elevation_m * 3.281, 1),
            "room_count": len(level.rooms),
            "head_count": len(heads_here),
            "pipe_count": len(pipes_here),
            "pipe_total_m": round(sum(p.length_m for p in pipes_here), 1),
            "pipe_total_ft": round(sum(p.length_m for p in pipes_here) * 3.281, 1),
        })
    return out


def _system_summary(design: Design) -> list[dict]:
    out = []
    for s in design.systems:
        h = s.hydraulic
        out.append({
            "id": s.id,
            "type": s.type,
            "supplies": s.supplies,
            "head_count": len(s.heads),
            "pipe_count": len(s.pipes),
            "pipe_total_m": round(sum(p.length_m for p in s.pipes), 1),
            "hanger_count": len(s.hangers),
            "riser_position_m": list(s.riser.position_m),
            "riser_size_in": s.riser.size_in,
            "fdc_type": s.riser.fdc_type,
            "hydraulic": {
                "design_area_sqft": h.design_area_sqft if h else None,
                "density_gpm_per_sqft": h.density_gpm_per_sqft if h else None,
                "required_flow_gpm": h.required_flow_gpm if h else None,
                "required_pressure_psi": h.required_pressure_psi if h else None,
                "supply_static_psi": h.supply_static_psi if h else None,
                "supply_residual_psi": h.supply_residual_psi if h else None,
                "demand_psi": h.demand_at_base_of_riser_psi if h else None,
                "safety_margin_psi": h.safety_margin_psi if h else None,
            } if h else None,
        })
    return out


def build_proposal_data(
    design: Design, bom: list[BomRow], labor: list[LaborRow],
    violations: list[dict] | None = None,
) -> dict[str, Any]:
    """Canonical proposal.json payload.

    This is what the web bid viewer reads. Stable API — new fields get
    added, existing ones never renamed.
    """
    bom_total = round(sum(r.extended_usd for r in bom), 2)
    labor_total = round(sum(r.extended_usd for r in labor), 2)
    subtotal = round(bom_total + labor_total, 2)
    permit = 3250.00  # Halo's typical included allowance
    taxes = round(subtotal * 0.072, 2)  # AZ rate approximation
    total = round(subtotal + permit + taxes, 2)

    return {
        "version": 1,
        "generated_at": date.today().isoformat(),
        "project": design.project.model_dump(),
        "building_summary": {
            "total_sqft": design.building.total_sqft,
            "construction_type": design.building.construction_type,
            "level_count": len(design.building.levels),
        },
        "levels": _level_summary(design),
        "systems": _system_summary(design),
        "scope_of_work": [
            "Perform fire sprinkler design, calculations, and material "
            "listings booklet for deferred submittals permit review",
            "Provide materials, equipment, & skilled labor to build the "
            "fully-functioning systems as permitted",
            "Garage dry systems + combination wet standpipe & sprinkler systems",
            "Schedule and complete inspections as required to close out permit",
            "Installed in accordance with NFPA and local AHJ code",
        ],
        "acknowledgements": [
            "Pricing valid for 10 days from the date of this proposal",
            "Scope begins 6\" above finished floor at the riser flange",
            "All underground by others",
            "Installation lead times to be coordinated no less than 6 weeks in advance",
            "Estimate assumes sufficient water flow; no fire pump",
            "Import, domestic, light wall, and threaded light wall NFPA piping acceptable",
            "A temporary standpipe will be installed when framed to the 4th floor",
            "Halo assumed a manual wet standpipe for this estimate",
            "Work schedule: 8.5 hours/day, 5 days/week",
        ],
        "inclusions": [
            "Pricing covers install based on documents provided for bidding",
            "Project management, travel, close-out docs, submittals included",
            "New piping, fittings, valves, sprinkler heads per bid documents",
            "Sales taxes included",
            "Scissor lifts, service vehicles, and small hand tools included",
            "White or chrome semi-recessed heads in finished ceilings",
            f"Permit fees included up to ${permit:,.2f}",
            "8 Rough-in mobilizations + 8 Trim mobilizations",
            "Wall-mount FDC at riser room",
        ],
        "exclusions": [
            "Pricing increases received after the date of this proposal",
            "Wiring, alarm, temp power, smoke & heat vents, FD access doors",
            "Pre-action, anti-freeze, hose valves, wet chemical systems",
            "Premium time, double time, prevailing wages",
            "Heads around equipment not shown on bid documents",
            "Paint, paint prep, refuse containers, haul-off",
            "3D modeling / BIM coordination",
            "Building signage, fire extinguishers & cabinets",
            "Cutting, patching, ceiling tile work",
            "Underground fire service line to building",
            "Galvanized steel for garage dry system + overhangs",
        ],
        "bom": [r.model_dump() for r in bom],
        "labor": [r.model_dump() for r in labor],
        "violations": violations or [],
        "pricing": {
            "materials_usd": bom_total,
            "labor_usd": labor_total,
            "permit_allowance_usd": permit,
            "taxes_usd": taxes,
            "subtotal_usd": subtotal,
            "total_usd": total,
        },
        "deliverables": {
            "sheet_set_pdf": "FP-0 cover, FP-H hydraulic placard, per-level FP-N plans, FP-R riser detail, FP-B BOM, FP-D details",
            "ifc_export": "IFC 4.x sprinkler subset for GC coordination model",
            "dxf_export": "AutoCAD DXF with AutoSprink-compatible layer names",
            "web_bid_viewer": "Interactive 3D model at /bid/{project_id}",
            "cut_sheets": "Manufacturer data sheets bundled PDF",
        },
    }


def write_proposal_files(
    data: dict[str, Any], out_dir: Path,
    design_payload: dict[str, Any] | None = None,
) -> dict[str, str]:
    """Emit proposal.json + HTML + PDF + XLSX. Returns file paths.

    ``design_payload`` is optional design.json content used by the
    HTML generator to draw per-level plan SVGs. If it's None, the
    HTML still renders — just without inline plans.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    paths: dict[str, str] = {}

    # JSON (the canonical artifact)
    json_path = out_dir / "proposal.json"
    json_path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")
    paths["json"] = str(json_path)

    # HTML — self-contained, embeds design.glb via <model-viewer>.
    # This is the artifact the VPS halo-fire client demo consumes.
    try:
        import importlib.util
        _spec = importlib.util.spec_from_file_location(
            "_halofire_proposal_html",
            str(Path(__file__).with_name("proposal_html.py")),
        )
        assert _spec is not None and _spec.loader is not None
        _mod = importlib.util.module_from_spec(_spec)
        _spec.loader.exec_module(_mod)
        write_proposal_html = _mod.write_proposal_html

        html_path = write_proposal_html(
            data, out_dir,
            design=design_payload,
            design_glb="design.glb",
        )
        paths["html"] = str(html_path)
    except Exception as e:
        paths["html_error"] = str(e)

    # Submittal sheet set — what the AHJ reviews
    try:
        import importlib.util as _ilu
        _sspec = _ilu.spec_from_file_location(
            "_halofire_submittal",
            str(Path(__file__).with_name("submittal.py")),
        )
        assert _sspec is not None and _sspec.loader is not None
        _smod = _ilu.module_from_spec(_sspec)
        _sspec.loader.exec_module(_smod)
        # The caller may pass design_payload; it's also used below.
        sub_path = _smod.write_submittal_pdf(
            data, out_dir, design=design_payload,
        )
        paths["submittal"] = str(sub_path)
    except Exception as e:
        paths["submittal_error"] = str(e)

    # Cut-sheet PDF bundle (one sheet per SKU, merged)
    try:
        import importlib.util as _ilu
        _cspec = _ilu.spec_from_file_location(
            "_halofire_cut_sheets",
            str(Path(__file__).with_name("cut_sheets.py")),
        )
        assert _cspec is not None and _cspec.loader is not None
        _cmod = _ilu.module_from_spec(_cspec)
        _cspec.loader.exec_module(_cmod)
        cs_res = _cmod.write_cut_sheet_bundle(
            data.get("bom") or [], out_dir,
            shared_library=out_dir.parent.parent / "cut_sheets_library",
        )
        paths["cut_sheets"] = cs_res.get("path", "")
        paths["cut_sheets_stubbed"] = str(
            len(cs_res.get("stubbed") or []),
        )
    except Exception as e:
        paths["cut_sheets_error"] = str(e)

    # PDF (reportlab)
    try:
        from reportlab.lib.pagesizes import LETTER
        from reportlab.lib.units import inch
        from reportlab.pdfgen import canvas

        pdf_path = out_dir / "proposal.pdf"
        c = canvas.Canvas(str(pdf_path), pagesize=LETTER)
        w, h = LETTER
        c.setFont("Helvetica-Bold", 16)
        c.drawString(0.75 * inch, h - 0.75 * inch, "HALO FIRE PROTECTION, LLC")
        c.setFont("Helvetica", 10)
        c.drawString(0.75 * inch, h - 1.0 * inch, data["project"]["name"])
        c.drawString(0.75 * inch, h - 1.2 * inch, data["project"]["address"])
        c.drawString(0.75 * inch, h - 1.4 * inch,
                     f"Proposal date: {data['generated_at']}")
        c.setFont("Helvetica-Bold", 18)
        c.drawString(0.75 * inch, h - 2.0 * inch,
                     f"Total price: ${data['pricing']['total_usd']:,.2f}")
        c.setFont("Helvetica", 10)
        y = h - 2.6 * inch
        c.drawString(0.75 * inch, y, "Scope of Work:")
        y -= 0.2 * inch
        for s in data["scope_of_work"]:
            c.drawString(1.0 * inch, y, f"• {s[:90]}")
            y -= 0.18 * inch
        y -= 0.2 * inch
        c.drawString(0.75 * inch, y, "Pricing breakdown:")
        y -= 0.2 * inch
        for k, v in data["pricing"].items():
            c.drawString(1.0 * inch, y, f"  {k:<25} ${v:>12,.2f}")
            y -= 0.18 * inch
        c.showPage()
        c.save()
        paths["pdf"] = str(pdf_path)
    except Exception as e:
        paths["pdf_error"] = str(e)

    # XLSX (openpyxl)
    try:
        from openpyxl import Workbook

        xlsx_path = out_dir / "proposal.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["HaloFire Proposal", data["project"]["name"]])
        ws.append(["Address", data["project"]["address"]])
        ws.append(["Date", data["generated_at"]])
        ws.append([""])
        ws.append(["Pricing"])
        for k, v in data["pricing"].items():
            ws.append([k, v])
        # BOM sheet
        ws_bom = wb.create_sheet("BOM")
        ws_bom.append(["SKU", "Description", "Qty", "Unit", "Unit $", "Extended $"])
        for r in data["bom"]:
            ws_bom.append([
                r["sku"], r["description"], r["qty"], r["unit"],
                r["unit_cost_usd"], r["extended_usd"],
            ])
        # Labor sheet
        ws_lab = wb.create_sheet("Labor")
        ws_lab.append(["Role", "Hours", "Rate $/hr", "Extended $"])
        for r in data["labor"]:
            ws_lab.append([r["role"], r["hours"], r["rate_usd_hr"], r["extended_usd"]])
        wb.save(xlsx_path)
        paths["xlsx"] = str(xlsx_path)
    except Exception as e:
        paths["xlsx_error"] = str(e)

    return paths


if __name__ == "__main__":
    print("proposal — call build_proposal_data(design, bom, labor)")
