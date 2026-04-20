"""Thin typed wrapper around the supplies.duckdb file.

Every bit of the pipeline that needs a price goes through this
module. There is no second source of truth.

    from pricing.db import open_db, price_for, apply_updates

    with open_db() as db:
        row = db.price_for("SM_Head_Pendant_Standard_K56")
        # -> PriceRow(unit_cost_usd=8.42, observed_at=..., stale=False)
"""
from __future__ import annotations

import hashlib
import json
from contextlib import contextmanager
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Iterator

try:
    import duckdb
except ImportError as e:  # pragma: no cover
    raise ImportError(
        "duckdb is required for pricing; `pip install duckdb`",
    ) from e


_HERE = Path(__file__).resolve().parent
_SCHEMA_SQL = (_HERE / "schema.sql").read_text(encoding="utf-8")
_DEFAULT_DB_PATH = _HERE / "supplies.duckdb"

# Staleness threshold — a price older than this is flagged. Bids built
# on stale prices are one of the top reasons Halo loses money.
STALE_DAYS = 60


@dataclass
class PriceRow:
    sku: str
    unit_cost_usd: float
    unit: str
    observed_at: datetime
    source: str
    confidence: float
    currency: str
    stale: bool  # true iff observed_at < now - STALE_DAYS


@dataclass
class PriceUpdate:
    """A single update produced by the sync agent."""

    sku: str
    unit_cost_usd: float
    unit: str = "ea"
    source: str = "manual"
    source_doc_sha256: str | None = None
    confidence: float = 1.0
    currency: str = "USD"

    def validate(self) -> list[str]:
        errs: list[str] = []
        if not self.sku or not isinstance(self.sku, str):
            errs.append("sku required")
        if self.unit_cost_usd is None or self.unit_cost_usd < 0:
            errs.append("unit_cost_usd must be >= 0")
        if self.unit not in {"ea", "ft", "m", "lb", "100ft", "each_100"}:
            errs.append(f"unit {self.unit!r} not in {{ea, ft, m, lb, 100ft, each_100}}")
        if not 0 <= self.confidence <= 1:
            errs.append("confidence must be in [0, 1]")
        return errs


@dataclass
class SyncRun:
    supplier_id: str | None = None
    source_url: str | None = None
    source_doc_sha256: str | None = None
    llm_model: str | None = None
    started_at: datetime = field(default_factory=datetime.utcnow)
    finished_at: datetime | None = None
    parts_touched: int = 0
    prices_added: int = 0
    status: str = "pending"
    error: str | None = None
    id: int | None = None


class SuppliesDB:
    """Session-scoped handle. Use via `open_db()`."""

    def __init__(self, path: Path | str) -> None:
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._con = duckdb.connect(str(self.path))
        self._bootstrap_schema()

    # ── lifecycle ──────────────────────────────────────────────

    def _bootstrap_schema(self) -> None:
        # DuckDB 1.x accepts a multi-statement script as a single
        # execute() call. Schema is ordered so sequences come
        # before the tables that DEFAULT on them.
        self._con.execute(_SCHEMA_SQL)

    def close(self) -> None:
        self._con.close()

    # ── suppliers / parts ─────────────────────────────────────

    def upsert_supplier(
        self,
        supplier_id: str,
        name: str,
        website: str | None = None,
        price_sheet_url: str | None = None,
        strategy: str | None = None,
        notes: str | None = None,
    ) -> None:
        self._con.execute(
            """
            INSERT OR REPLACE INTO suppliers
              (id, name, website, price_sheet_url, strategy, notes)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            [supplier_id, name, website, price_sheet_url, strategy, notes],
        )

    def upsert_part(self, part: dict[str, Any]) -> None:
        """Idempotent part insert. `part` keys match the schema columns."""
        cols = [
            "sku", "name", "category", "mounting", "manufacturer",
            "supplier_id", "model", "dim_l_cm", "dim_d_cm", "dim_h_cm",
            "pipe_size_in", "k_factor", "temp_rating_f", "response",
            "connection", "finish", "nfpa_paint_hex", "open_source_glb",
            "discontinued", "notes",
        ]
        values = [part.get(c) for c in cols]
        placeholders = ", ".join(["?"] * len(cols))
        assignments = ", ".join(f"{c}=excluded.{c}" for c in cols if c != "sku")
        # DuckDB's parameter binder is strict about `CURRENT_TIMESTAMP`
        # appearing mid-statement when `?` placeholders are in use —
        # commit the timestamp in a follow-up UPDATE.
        self._con.execute(
            f"""
            INSERT INTO parts ({", ".join(cols)})
            VALUES ({placeholders})
            ON CONFLICT (sku) DO UPDATE SET
              {assignments}
            """,
            values,
        )
        self._con.execute(
            "UPDATE parts SET updated_at = CURRENT_TIMESTAMP WHERE sku = ?",
            [part["sku"]],
        )

    # ── prices ────────────────────────────────────────────────

    def add_price(self, update: PriceUpdate) -> int:
        errs = update.validate()
        if errs:
            raise ValueError(f"invalid PriceUpdate for {update.sku}: {errs}")
        # Ensure the part exists — append-only prices pointing at a
        # non-existent SKU is the kind of silent corruption we will
        # not allow.
        exists = self._con.execute(
            "SELECT 1 FROM parts WHERE sku = ? LIMIT 1", [update.sku],
        ).fetchone()
        if exists is None:
            raise KeyError(f"sku {update.sku!r} is not in parts — seed it first")
        row = self._con.execute(
            """
            INSERT INTO prices
              (sku, unit_cost_usd, unit, source,
               source_doc_sha256, confidence, currency)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            RETURNING id
            """,
            [
                update.sku,
                float(update.unit_cost_usd),
                update.unit,
                update.source,
                update.source_doc_sha256,
                float(update.confidence),
                update.currency,
            ],
        ).fetchone()
        return int(row[0]) if row else 0

    def apply_updates(
        self, updates: Iterable[PriceUpdate],
    ) -> tuple[int, list[str]]:
        """Validate + commit a batch. Returns (accepted, errors)."""
        accepted = 0
        errs: list[str] = []
        for u in updates:
            try:
                self.add_price(u)
                accepted += 1
            except Exception as e:  # noqa: BLE001
                errs.append(f"{u.sku}: {e}")
        return accepted, errs

    def price_for(
        self,
        sku: str,
        as_of: datetime | None = None,
    ) -> PriceRow | None:
        """Return the price observation to use for a BOM line.

        `as_of` pins the query to a historical moment (rebuilding an
        old bid). Default = now.
        """
        cutoff = as_of or datetime.utcnow()
        row = self._con.execute(
            """
            SELECT sku, unit_cost_usd, unit, observed_at, source,
                   confidence, currency
            FROM prices
            WHERE sku = ? AND observed_at <= ?
            ORDER BY observed_at DESC
            LIMIT 1
            """,
            [sku, cutoff],
        ).fetchone()
        if row is None:
            return None
        observed_at = row[3] if isinstance(row[3], datetime) else datetime.fromisoformat(str(row[3]))
        stale = observed_at < cutoff - timedelta(days=STALE_DAYS)
        return PriceRow(
            sku=row[0],
            unit_cost_usd=float(row[1]),
            unit=row[2],
            observed_at=observed_at,
            source=row[4],
            confidence=float(row[5]),
            currency=row[6],
            stale=stale,
        )

    def stale_skus(self) -> list[tuple[str, int]]:
        """All SKUs whose latest price is older than STALE_DAYS (or missing)."""
        rows = self._con.execute(
            "SELECT sku, COALESCE(days_stale, -1) FROM stale_skus "
            "ORDER BY days_stale DESC NULLS FIRST",
        ).fetchall()
        return [(r[0], int(r[1])) for r in rows]

    def all_parts(self) -> list[dict[str, Any]]:
        cols = [
            "sku", "name", "category", "manufacturer", "model",
            "pipe_size_in", "k_factor", "finish", "discontinued",
        ]
        rows = self._con.execute(
            f"SELECT {', '.join(cols)} FROM parts ORDER BY category, sku",
        ).fetchall()
        return [dict(zip(cols, r)) for r in rows]

    def part_count(self) -> int:
        return int(self._con.execute("SELECT COUNT(*) FROM parts").fetchone()[0])

    # ── sync runs ─────────────────────────────────────────────

    def start_sync_run(self, run: SyncRun) -> int:
        row = self._con.execute(
            """
            INSERT INTO sync_runs
              (started_at, supplier_id, source_url,
               source_doc_sha256, llm_model, status)
            VALUES (?, ?, ?, ?, ?, 'pending')
            RETURNING id
            """,
            [
                run.started_at,
                run.supplier_id,
                run.source_url,
                run.source_doc_sha256,
                run.llm_model,
            ],
        ).fetchone()
        run.id = int(row[0])
        return run.id

    def finish_sync_run(
        self,
        run_id: int,
        parts_touched: int,
        prices_added: int,
        status: str,
        error: str | None = None,
    ) -> None:
        self._con.execute(
            """
            UPDATE sync_runs
            SET finished_at = CURRENT_TIMESTAMP,
                parts_touched = ?,
                prices_added = ?,
                status = ?,
                error = ?
            WHERE id = ?
            """,
            [parts_touched, prices_added, status, error, run_id],
        )

    def recent_sync_runs(self, limit: int = 20) -> list[dict[str, Any]]:
        cols = [
            "id", "started_at", "finished_at", "supplier_id",
            "source_url", "parts_touched", "prices_added",
            "status", "error", "llm_model",
        ]
        rows = self._con.execute(
            f"SELECT {', '.join(cols)} FROM sync_runs "
            "ORDER BY started_at DESC LIMIT ?",
            [limit],
        ).fetchall()
        return [dict(zip(cols, r)) for r in rows]

    # ── Excel bridge ──────────────────────────────────────────

    def export_xlsx(self, out_path: Path | str) -> Path:
        """Dump parts + latest_prices + sync_runs to an XLSX snapshot."""
        out = Path(out_path)
        out.parent.mkdir(parents=True, exist_ok=True)
        # DuckDB's native Excel writer
        self._con.execute(
            "INSTALL excel; LOAD excel;"
        )
        tmp_parts = self._con.execute(
            "SELECT * FROM parts ORDER BY category, sku",
        ).fetchdf()
        tmp_prices = self._con.execute(
            "SELECT * FROM latest_prices ORDER BY sku",
        ).fetchdf()
        tmp_runs = self._con.execute(
            "SELECT * FROM sync_runs ORDER BY started_at DESC",
        ).fetchdf()
        # openpyxl for multi-sheet; DuckDB EXPORT DATABASE is
        # per-table only.
        from openpyxl import Workbook

        def _to_xlsx(ws_obj, df) -> None:  # type: ignore[no-untyped-def]
            # pandas may emit pd.NA / NaT for nullable columns — openpyxl
            # rejects those. Normalize to None.
            ws_obj.append(list(df.columns))
            import pandas as _pd  # local import to keep module lightweight

            for row in df.itertuples(index=False, name=None):
                ws_obj.append(
                    [
                        None if _pd.isna(v) else (v.isoformat() if hasattr(v, "isoformat") else v)
                        for v in row
                    ],
                )

        wb = Workbook()
        ws = wb.active
        ws.title = "parts"
        _to_xlsx(ws, tmp_parts)
        _to_xlsx(wb.create_sheet("latest_prices"), tmp_prices)
        _to_xlsx(wb.create_sheet("sync_runs"), tmp_runs)
        wb.save(out)
        return out

    # ── raw ──────────────────────────────────────────────────

    def con(self) -> duckdb.DuckDBPyConnection:
        """Escape hatch for ad-hoc SQL. Avoid; most flows have a method."""
        return self._con


# ── helpers ──────────────────────────────────────────────────

def sha256_of(path: Path | str) -> str:
    h = hashlib.sha256()
    with Path(path).open("rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


@contextmanager
def open_db(path: Path | str | None = None) -> Iterator[SuppliesDB]:
    db = SuppliesDB(path or _DEFAULT_DB_PATH)
    try:
        yield db
    finally:
        db.close()


# ── top-level shortcuts ─────────────────────────────────────

def price_for(sku: str, as_of: datetime | None = None) -> PriceRow | None:
    with open_db() as db:
        return db.price_for(sku, as_of=as_of)


def apply_updates(
    updates: Iterable[PriceUpdate],
) -> tuple[int, list[str]]:
    with open_db() as db:
        return db.apply_updates(updates)


def dump_json(out_path: Path | str) -> Path:
    """Emit every part + latest price as a single JSON — for audit."""
    out = Path(out_path)
    with open_db() as db:
        parts = db.all_parts()
        priced = []
        for p in parts:
            row = db.price_for(p["sku"])
            priced.append(
                {
                    **p,
                    "price": (
                        {
                            "unit_cost_usd": row.unit_cost_usd,
                            "unit": row.unit,
                            "observed_at": row.observed_at.isoformat(),
                            "stale": row.stale,
                            "source": row.source,
                        }
                        if row
                        else None
                    ),
                },
            )
    out.write_text(json.dumps(priced, indent=2, default=str), encoding="utf-8")
    return out


__all__ = [
    "SuppliesDB",
    "PriceRow",
    "PriceUpdate",
    "SyncRun",
    "STALE_DAYS",
    "sha256_of",
    "open_db",
    "price_for",
    "apply_updates",
    "dump_json",
]
