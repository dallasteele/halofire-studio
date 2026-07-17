"""Copy-only additive SQLite migration for the HaloFire operational spine.

The bid corpus is an answer key. This tool always copies it before applying
namespaced platform tables and records the source digest in a receipt.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import sqlite3
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

MIGRATION_ID = "platform-spine-0001"

UP_SQL = """
CREATE TABLE IF NOT EXISTS platform_migrations (id TEXT PRIMARY KEY, source_sha256 TEXT NOT NULL, applied_at TEXT NOT NULL);
CREATE TABLE IF NOT EXISTS platform_customers (id TEXT PRIMARY KEY, display_name TEXT NOT NULL, source_table TEXT NOT NULL, source_row_id INTEGER NOT NULL, source_sha256 TEXT NOT NULL, created_at TEXT NOT NULL, UNIQUE(source_table, source_row_id));
CREATE TABLE IF NOT EXISTS platform_sites (id TEXT PRIMARY KEY, customer_id TEXT, display_name TEXT NOT NULL, address TEXT, source_table TEXT NOT NULL, source_row_id INTEGER NOT NULL, source_sha256 TEXT NOT NULL, created_at TEXT NOT NULL, UNIQUE(source_table, source_row_id));
CREATE TABLE IF NOT EXISTS platform_jobs (id TEXT PRIMARY KEY, customer_id TEXT, site_id TEXT, display_name TEXT NOT NULL, status TEXT NOT NULL, source_table TEXT NOT NULL, source_row_id INTEGER NOT NULL, source_sha256 TEXT NOT NULL, created_at TEXT NOT NULL, UNIQUE(source_table, source_row_id));
CREATE TABLE IF NOT EXISTS platform_employees (id TEXT PRIMARY KEY, display_name TEXT NOT NULL, email TEXT, role TEXT NOT NULL, active INTEGER NOT NULL, source_table TEXT NOT NULL, source_row_id INTEGER NOT NULL, source_sha256 TEXT NOT NULL, created_at TEXT NOT NULL, UNIQUE(source_table, source_row_id));
CREATE TABLE IF NOT EXISTS platform_documents (id TEXT PRIMARY KEY, job_id TEXT, locator_path TEXT NOT NULL, sha256 TEXT, kind TEXT NOT NULL, source_table TEXT NOT NULL, source_row_id INTEGER NOT NULL, source_sha256 TEXT NOT NULL, created_at TEXT NOT NULL, UNIQUE(source_table, source_row_id));
CREATE TABLE IF NOT EXISTS platform_audit_events (id TEXT PRIMARY KEY, actor_employee_id TEXT, action TEXT NOT NULL, entity_kind TEXT NOT NULL, entity_id TEXT NOT NULL, payload_sha256 TEXT, created_at TEXT NOT NULL);
CREATE TABLE IF NOT EXISTS platform_review_items (id TEXT PRIMARY KEY, job_id TEXT, module_id TEXT NOT NULL, kind TEXT NOT NULL, severity TEXT NOT NULL CHECK(severity IN ('advisory','hard')), status TEXT NOT NULL CHECK(status IN ('open','resolved','dismissed')), evidence_sha256 TEXT, created_at TEXT NOT NULL, updated_at TEXT NOT NULL);
CREATE TABLE IF NOT EXISTS platform_module_registry (module_id TEXT PRIMARY KEY, display_name TEXT NOT NULL, nav_path TEXT NOT NULL, enabled INTEGER NOT NULL, created_at TEXT NOT NULL, updated_at TEXT NOT NULL);
CREATE TABLE IF NOT EXISTS platform_autobid_links (source_table TEXT NOT NULL, source_row_id INTEGER NOT NULL, platform_job_id TEXT NOT NULL REFERENCES platform_jobs(id), source_sha256 TEXT NOT NULL, created_at TEXT NOT NULL, PRIMARY KEY(source_table, source_row_id));
"""

def digest(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""): h.update(chunk)
    return h.hexdigest()

def workbook_value(value: Any) -> str:
    """Return a stable cell value without changing the source workbook."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()

def read_employee_roster(workbook: Path) -> tuple[list[dict[str, str | int | None]], str]:
    """Read the authoritative Halo contact list into fail-closed import rows.

    The roster contains contact data, not employment state or permissions.
    Imported employees are therefore disabled and unassigned until M5 creates a
    source-backed authorization mapping.
    """
    if not workbook.is_file():
        raise ValueError("PLATFORM_EMPLOYEE_ROSTER_MISSING")
    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError as exc:
        raise ValueError("PLATFORM_EMPLOYEE_ROSTER_READER_UNAVAILABLE") from exc
    book = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    try:
        if "Sheet1" not in book.sheetnames:
            raise ValueError("PLATFORM_EMPLOYEE_ROSTER_SHEET_MISSING")
        sheet = book["Sheet1"]
        headers = {workbook_value(sheet.cell(1, column).value).casefold(): column for column in range(1, sheet.max_column + 1)}
        required = ("employee id", "first name", "last name", "email work")
        if any(header not in headers for header in required):
            raise ValueError("PLATFORM_EMPLOYEE_ROSTER_HEADERS_UNSUPPORTED")
        records: list[dict[str, str | int | None]] = []
        seen_ids: set[str] = set()
        for row_number in range(2, sheet.max_row + 1):
            employee_id = workbook_value(sheet.cell(row_number, headers["employee id"]).value)
            first_name = workbook_value(sheet.cell(row_number, headers["first name"]).value)
            last_name = workbook_value(sheet.cell(row_number, headers["last name"]).value)
            work_email = workbook_value(sheet.cell(row_number, headers["email work"]).value)
            if not any((employee_id, first_name, last_name, work_email)):
                continue
            if not employee_id or not first_name or not last_name:
                raise ValueError("PLATFORM_EMPLOYEE_ROSTER_ROW_INCOMPLETE")
            if employee_id in seen_ids:
                raise ValueError("PLATFORM_EMPLOYEE_ROSTER_DUPLICATE_ID")
            if work_email and (work_email.count("@") != 1 or "." not in work_email.rsplit("@", 1)[1]):
                raise ValueError("PLATFORM_EMPLOYEE_ROSTER_EMAIL_INVALID")
            seen_ids.add(employee_id)
            records.append({"employee_id": employee_id, "display_name": f"{first_name} {last_name}", "email": work_email or None, "row_number": row_number})
        if not records:
            raise ValueError("PLATFORM_EMPLOYEE_ROSTER_EMPTY")
        return records, digest(workbook)
    finally:
        book.close()

def apply(source: Path, target: Path, receipt: Path, employees_workbook: Path | None = None) -> dict[str, object]:
    if source.resolve() == target.resolve(): raise ValueError("PLATFORM_MIGRATION_IN_PLACE_REJECTED")
    if not source.is_file(): raise ValueError("PLATFORM_MIGRATION_SOURCE_MISSING")
    employee_records: list[dict[str, str | int | None]] = []
    employee_source_sha: str | None = None
    if employees_workbook is not None:
        employee_records, employee_source_sha = read_employee_roster(employees_workbook)
    source_sha = digest(source); target.parent.mkdir(parents=True, exist_ok=True); shutil.copy2(source, target)
    now = datetime.now(UTC).isoformat(); conn = sqlite3.connect(target)
    try:
        conn.executescript(UP_SQL)
        for row_id, name in conn.execute("SELECT id, name FROM builders WHERE name IS NOT NULL ORDER BY id"):
            conn.execute("INSERT OR IGNORE INTO platform_customers VALUES (?,?,?,?,?,?)", (f"autobid:builders:{row_id}", name, "builders", row_id, source_sha, now))
        for row_id, builder_id, name in conn.execute("SELECT id, builder_id, name FROM jobs WHERE name IS NOT NULL ORDER BY id"):
            customer_id = f"autobid:builders:{builder_id}" if builder_id is not None else None
            conn.execute("INSERT OR IGNORE INTO platform_jobs VALUES (?,?,?,?,?,?,?,?,?)", (f"autobid:jobs:{row_id}", customer_id, None, name, "bidding", "jobs", row_id, source_sha, now))
        for row_id, job_id, rel_path, filename in conn.execute("SELECT id, job_id, rel_path, filename FROM documents WHERE rel_path IS NOT NULL ORDER BY id"):
            conn.execute("INSERT OR IGNORE INTO platform_documents VALUES (?,?,?,?,?,?,?,?,?)", (f"autobid:documents:{row_id}", f"autobid:jobs:{job_id}" if job_id is not None else None, rel_path, None, filename or "document", "documents", row_id, source_sha, now))
        for record in employee_records:
            conn.execute(
                "INSERT OR IGNORE INTO platform_employees VALUES (?,?,?,?,?,?,?,?,?)",
                (f"halo-roster:sheet1:{record['employee_id']}", record["display_name"], record["email"], "unassigned", 0, "halo_contact_list_sheet1", record["row_number"], employee_source_sha, now),
            )
        for table, query in {
            "intake_bids": "SELECT i.id, j.id FROM intake_bids i JOIN builders b ON b.name=i.builder JOIN jobs j ON j.builder_id=b.id AND j.name=i.job",
            "ready_bids": "SELECT r.id, d.job_id FROM ready_bids r JOIN documents d ON d.id=r.document_id WHERE d.job_id IS NOT NULL",
            "bid_summary": "SELECT s.id, d.job_id FROM bid_summary s JOIN documents d ON d.id=s.document_id WHERE d.job_id IS NOT NULL",
        }.items():
            for source_row_id, job_id in conn.execute(query):
                conn.execute("INSERT OR IGNORE INTO platform_autobid_links VALUES (?,?,?,?,?)", (table, source_row_id, f"autobid:jobs:{job_id}", source_sha, now))
        conn.execute("INSERT OR REPLACE INTO platform_migrations VALUES (?,?,?)", (MIGRATION_ID, source_sha, now)); conn.commit()
        counts = {name: conn.execute(f"SELECT COUNT(*) FROM {name}").fetchone()[0] for name in ("platform_customers", "platform_jobs", "platform_documents", "platform_sites", "platform_employees", "platform_autobid_links")}
    finally: conn.close()
    result = {"migration": MIGRATION_ID, "source": str(source), "target": str(target), "source_sha256": source_sha, "target_sha256": digest(target), "employee_source": None if employees_workbook is None else {"path": str(employees_workbook), "sha256": employee_source_sha, "imported": len(employee_records), "authorization_state": "unassigned_disabled"}, "counts": counts, "source_writable": False}
    receipt.write_text(json.dumps(result, indent=2) + "\n", encoding="utf-8"); return result

def rollback(target: Path) -> dict[str, object]:
    if not target.is_file(): raise ValueError("PLATFORM_MIGRATION_TARGET_MISSING")
    conn = sqlite3.connect(target)
    try:
        applied = conn.execute("SELECT source_sha256 FROM platform_migrations WHERE id=?", (MIGRATION_ID,)).fetchone()
        if applied is None: raise ValueError("PLATFORM_MIGRATION_NOT_APPLIED")
        for table in ("platform_autobid_links", "platform_module_registry", "platform_review_items", "platform_audit_events", "platform_documents", "platform_employees", "platform_jobs", "platform_sites", "platform_customers", "platform_migrations"):
            conn.execute(f"DROP TABLE {table}")
        conn.commit()
        remaining = conn.execute("SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name LIKE 'platform_%'").fetchone()[0]
    finally: conn.close()
    return {"migration": MIGRATION_ID, "target": str(target), "rolled_back": True, "remaining_platform_tables": remaining}

def main() -> None:
    p = argparse.ArgumentParser(); p.add_argument("--source", type=Path); p.add_argument("--target", type=Path); p.add_argument("--receipt", type=Path); p.add_argument("--employees-workbook", type=Path); p.add_argument("--rollback-target", type=Path); args = p.parse_args()
    if args.rollback_target: print(json.dumps(rollback(args.rollback_target))); return
    if not (args.source and args.target and args.receipt): p.error("--source, --target, and --receipt are required unless --rollback-target is supplied")
    print(json.dumps(apply(args.source, args.target, args.receipt, args.employees_workbook)))
if __name__ == "__main__": main()
